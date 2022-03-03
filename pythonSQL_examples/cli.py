import click
from openpyxl import load_workbook
import mysql.connector
from mysql.connector import Error
# variables de entorno
from dotenv import load_dotenv
import os 
# log
import logging
#excel


logging.basicConfig(
    level=logging.DEBUG,
    format="{asctime} {levelname:<8} {message}",
    style= '{',
    filename= 'mylog.log',
    filemode= 'w'
)

load_dotenv()

#inserta datos a traves de un excel
@click.command()
@click.option('--filename', type=click.Path(exists=True), help='ruta del archivo')
def file(filename):
    # repite un saludo a alguiuen determinado numero de veces
    #hojas para manipular el excel
    workbook = load_workbook(filename='./MOCK_DATA.xlsx')
    sheet = workbook.active
    try:
        conexion = mysql.connector.connect(
            host = os.getenv('host'),
            database = os.getenv('database'),
            user = os.getenv('user'),
            password = os.getenv('password') 
        )

        if conexion.is_connected():
            db_info = conexion.get_server_info()
            logging.info(f'conexion a la DB establecida: {db_info}')

            # cursor nos sirve para ejecutar instrucciones sql
            cursor = conexion.cursor()
            cursor.execute("select database();")
            record = cursor.fetchone()
            logging.info(f"You're connected to database: {record}" )


            # insertar datos al sql desde datos obtenidos del excel
            for row in sheet.iter_rows(min_row=2, max_row=12, min_col=1, max_col=6, values_only=True):
                cursor.callproc('insertar_cliente', args=(row[1], row[2], row[3], row[4], row[5]))
                conexion.commit()
                logging.info("datos insertados")
    except Error as e:
        logging.error(f'error al conectar a la db: {e}')
    finally:
        if conexion.is_connected():
            cursor.close()
            conexion.close()
            logging.info('la conexion a la db se ha cerrado')
            click.echo('datos insertados en la DB')


file()