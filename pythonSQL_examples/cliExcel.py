from datatest import validate
import click
from openpyxl import load_workbook
# log
import logging
from conexion import ConexionDB
from logConf import *

c = ConexionDB()

#inserta datos a traves de un excel
@click.command()
@click.option('--filename', type=click.Path(exists=True), help='ruta del archivo')
def file(filename):
    #hojas para manipular el excel
    workbook = load_workbook(filename='./MOCK_DATA.xlsx')
    sheet = workbook.active
    
    # cursor
    cursor = c.conexion.cursor()

    # insertar datos al sql desde datos obtenidos del excel
    for row in sheet.iter_rows(min_row=2, max_row=11, min_col=1, max_col=6, values_only=True):
        validate(row[1], str) # first_name
        validate(row[2], str) # last_name
        validate(row[3], str) # email
        validate(row[4], str) # gender
        validate(row[5], str) # ip_address
        cursor.callproc('insertar_cliente', args=(row[1], row[2], row[3], row[4], row[5]))
        c.conexion.commit()
        logging.info("datos insertados")
    click.echo('datos insertados')


file()
